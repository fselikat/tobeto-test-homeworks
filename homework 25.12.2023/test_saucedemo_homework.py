from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
import pytest
import openpyxl
from constants import globalconstants as gc

class Test_DemoClass:
    #prefix => test_ 
    def setup_method(self): #her test başlangıcında çalışacak fonk
        self.driver = webdriver.Chrome()
        self.driver.get(gc.BASE_URL)
        self.driver.maximize_window() 

    def teardown_method(self): # her testinin bitiminde çalışacak fonk
        self.driver.quit()

    def getData1():
        excel = openpyxl.load_workbook(gc.TEST_INPUTS_XLSX)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,5):
            add = sheet.cell(i,1).value
            remove = sheet.cell(i,2).value
            data.append((add,remove))

        return data
    def getData2():
        excel = openpyxl.load_workbook(gc.TEST_INPUTS_XLSX)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,5):
            username = sheet.cell(i,4).value
            password = sheet.cell(i,5).value
            data.append((username, password))

        return data
    def getData3():
        excel = openpyxl.load_workbook(gc.TEST_INPUTS_XLSX)
        sheet = excel["Sayfa1"] #hangi sayfada çalışacağımı gösteriyorum
        rows = sheet.max_row #kaçıncı satıra kadar veri var?
        data = []
        for i in range(2,5):
            add = sheet.cell(i,7).value
            data.append((add))

        return data

    @pytest.mark.parametrize("add,remove",getData1())
    def test_error_user(self,add,remove):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(gc.ERROR_USER)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(gc.PASSWORD)
        loginButton = self.driver.find_element(By.ID, gc.LOGIN_BUTTON_ID)
        loginButton.click()
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,add)))
        addToCart.click()
        removeFromCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,remove)))
        removeFromCart.click()
        assert removeFromCart.text =="Remove"



    @pytest.mark.parametrize("username,password",getData2())
    def test_valid_login(self,username, password):
        main_page_url = self.driver.current_url
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(username)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.PASSWORD_ID)))
        passwordInput.send_keys(password)
        loginButton = self.driver.find_element(By.ID, gc.LOGIN_BUTTON_ID)
        loginButton.click()
        menuButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, gc.MENUBUTTON_ID)))
        menuButton.click()
        logoutLink = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, gc.LOGOUT_XPATH)))
        logoutLink.click()
        sleep(5)
        logout_url=self.driver.current_url
        assert main_page_url == logout_url

    @pytest.mark.parametrize("add",getData3())
    def test_addToCart(self,add):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,gc.USERNAME_ID)))
        usernameInput.send_keys(gc.STANDARD_USER)
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID, gc.PASSWORD_ID)))
        passwordInput.send_keys(gc.PASSWORD)
        loginButton = self.driver.find_element(By.ID, gc.LOGIN_BUTTON_ID)
        loginButton.click()
        self.driver.execute_script("window.scrollTo(0,500)") #penceremin scrolunu aşağıya indiriyor,verdiğim koordinatlarla birlikte
        addToCart = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.NAME, add)))
        addToCart.click()
        cartSize =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH, gc.CART_SIZE_XPATH)))
        assert cartSize.text =="1"
